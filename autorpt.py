#!/usr/bin/python3
"""autorpt.py - enforce consistent, dependable workflow for engagement note-taking and report writing"""

import blessings
import configparser
import csv
from cvss import CVSS3
import datetime
import getopt
from glob import glob
import json
import numpy as np
import os
import openpyxl
import pandas as pd
from pathlib import Path
import random
import re
import shutil
import subprocess
import sys
import time
import yaml

def helper():
    """The CLI help output"""
    print(f"{term.bold_bright_blue}USAGE:{term.normal}")
    print(f"autorpt.py [option]\n")
    print(f"{term.bold_bright_blue}Where option is one of:{term.normal}")
    print(f"help | startup | vuln | ports | sitrep [message] | finalize ")
    print(f"add [IP Address] | settings | active | whathaveidone\n")
    print(f"{term.bold_bright_blue}OPTIONS:{term.normal}")
    print(f'  {term.bright_blue}help{term.normal}      Display this listing of usage and examples.')
    print(f'  {term.bright_blue}startup{term.normal}   Create a clean working directory for a new engagement.')
    print(f'  {term.bright_blue}add{term.normal}       Add a newly discovered IP address to target.md and create its markdown file.')
    print(f'  {term.bright_blue}vuln{term.normal}      Record a confirmed vulnerability with CVSS scoring and MITRE ATT&CK attributes.')
    print(f'  {term.bright_blue}ports{term.normal}     (AutoRecon specific) Quick display of all open ports per target.')
    print(f'  {term.bright_blue}sitrep{term.normal}    Record a status update of your current progress or display the menu.')
    print(f'  {term.bright_blue}finalize{term.normal}  Compile markdown files into a desired output file format.')
    print(f'  {term.bright_blue}settings{term.normal}  Configuration settings.')
    print(f'  {term.bright_blue}active{term.normal}    Display the active engagement name and path.')

    print(f"\n{term.bold_bright_blue}EXAMPLES:{term.normal}")
    print("When you are ready to start an exam or training:")
    print(f"    {term.bright_blue}autorpt.py startup{term.normal}")
    print("Log a verified vulnerability:")
    print(f"    {term.bright_blue}autorpt.py vuln{term.normal}")
    print("Display vulnerability list:")
    print(f"    {term.bright_blue}autorpt.py vuln list{term.normal}")
    print("Log your current status:")
    print(f"    {term.bright_blue}autorpt.py sitrep Stuck trying to exploit system X:8001/login.php via SQLi.  May be a rabbit trail.{term.normal}")
    print("...Or use the menu system:")
    print(f"    {term.bright_blue}autorpt.py sitrep{term.normal}")
    print("Display the sitrep log:")
    print(f"    {term.bright_blue}autorpt.py sitrep list{term.normal}")
    print("After AutoRecon completes, display the ports:")
    print(f"    {term.bright_blue}autorpt.py ports{term.normal}")
    print("Compile the markdown into a polished report document")
    print(f"    {term.bright_blue}autorpt.py finalize{term.normal}")
    sys.exit(1)

def banner():
    """Display required ASCII art and random motto"""
    msg = ""
    mottos = ['Train like you PenTest',
            'Persistently consistent', 
            'We PenTest like we train',
            'Consistency is the key',
            'Train like you exam like you PenTest', 
            'Tag your work', 
            'Documentation is never perfect,\nit simply runs out of time',
            'Consistent, dependable, and improving']
    max_size = len(mottos) - 1
    random_message_id = random.randint(0, max_size)
    msg = mottos[random_message_id]
    print(f'')
    print(f'  ▄▄▄· ▄• ▄▌▄▄▄▄▄      ▄▄▄   ▄▄▄·▄▄▄▄▄ ')
    print(f' ▐█ ▀█ █▪██▌•██  ▪     ▀▄ █·▐█ ▄█•██   ')
    print(f' ▄█▀▀█ █▌▐█▌ ▐█.▪ ▄█▀▄ ▐▀▀▄  ██▀· ▐█.▪ ')
    print(f' ▐█ ▪▐▌▐█▄█▌ ▐█▌·▐█▌.▐▌▐█•█▌▐█▪·• ▐█▌· ')
    print(f'  ▀  ▀  ▀▀▀  ▀▀▀  ▀█▄▀▪.▀  ▀.▀    ▀▀▀  ')
    print(f'{term.bright_blue}{msg.center(40)}{term.normal}\n')
    
def clearScreen():
    _ = subprocess.call('clear' if os.name == 'posix' else 'cls')

def colorHeader(msg):
    print(f"\n{term.bold_bright_blue}{msg}{term.normal}\n")

def colorSubHeading(msg):
    print(f"{term.bold_bright_blue}{msg}{term.normal}")

def colorMenuItem(msg):
    print(f"  {term.yellow}{msg}{term.normal}")

def colorList(msg):
    print(f"{term.bold_bright}{msg}{term.normal}")

def colorDebug(msg):
    print(f"{term.on_yellow}{term.black}[d]{term.normal}  {term.yellow}{msg}{term.normal}")

def colorTableHeader(msg):
    print(f"{term.on_blue_underline_bold}{term.bright_white}{msg}{term.normal}")

# OBSOLETE >>>> 
# Pending migration
def colorVerification(field, msg):
    colorVerify(field, msg)

def colorVerificationPass(field, msg):
    colorPass(field, msg)

def colorVerificationFail(field, msg):
    colorFail(field, msg)
# OBSOLETE <<<<

def colorVerify(field, msg):
    print(f'{term.red}{field}{term.normal}  {term.bold_bright}{msg}{term.normal}')

def colorWarn(msg):
    print(f"{term.bold_on_bright_yellow}[W]{term.normal}  {term.bold_bright_yellow}{msg}{term.normal}")

def colorPass(field, msg):
    print(f'{term.bold_on_bright_green}[P]{term.normal}  {term.bold_bright_green}{field} {msg}{term.normal}')

def colorFail(field, msg):
    print(f'{term.bold_on_bright_red}[E]{term.normal}  {term.bold_bright_red}{field} {msg}{term.normal}')

def colorNotice(msg):
    print(f"{term.bright}{msg}{term.normal}")

def getCvss3Score():
    """Menu prompting for the vulnerability CVSS scoring"""
    colorHeader("CVSS 3 Scoring")
    print("Do you know the Overall CVSS v3 Score? [Y|N]")
    picker = str(input(">  ")).upper()
    if "Y" == picker:
        print("What is the score?")
        cvssScore = float(input(">  "))
        if cvssScore <= 3.9:
            cvssSeverity = 'Low'
        elif cvssScore >= 4.0 and cvssScore <= 6.9:
            cvssSeverity = 'Medium'
        elif cvssScore >= 7.0 and cvssScore <= 8.9:
            cvssSeverity = 'High'
        elif cvssScore >= 9.0 and cvssScore <= 10.0:
            cvssSeverity = 'Critical'
        else:
            colorNotice('The score must be between 0.1 and 10.0.')
            getCvss3Score
        print("If known, paste the CVSS Vector string here or hit Enter to skip (eg. AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:N ")
        cvssVector = str(input('>  ')).upper()
        if 'NONE' == cvssVector or '' == cvssVector:
            cvssVector = ''
        else:
            cvssVector = "CVSS:3.0/" + cvssVector
        returnStr = [cvssSeverity, str(cvssScore), cvssVector]
    else:
        colorSubHeading('Base Score Metrics')
        colorSubHeading('Exploit')
        cvssAVs = {'Network': 'AV:N', 
                    'Adjacent Network': 'AV:A', 
                    'Local': 'AV:L', 
                    'Physical': 'AV:P'}
        aV = cvssAVs[getCvssMetricValue(cvssAVs, "Attack Vector")]
        cvssACs = {'Low': 'AC:L', 'High': 'AC:H'}
        aC = cvssACs[getCvssMetricValue(cvssACs, "Attack Complexity")]
        cvssPRs = {'None': 'PR:N', 'Low': 'PR:L', 'High': 'PR:H'}
        pR = cvssPRs[getCvssMetricValue(cvssPRs, "Privileges Required")]
        cvssUIs = {'None': 'UI:N', 'Required': 'UI:R'}
        uI = cvssUIs[getCvssMetricValue(cvssUIs, "User Interaction")]
        cvssSs = {'Unchanged': 'S:U', 'Changed': 'S:C'}
        s = cvssSs[getCvssMetricValue(cvssSs, "Scope")]
        
        colorSubHeading('Impact')
        cvssCs = {'None': 'C:N', 'Low': 'C:L', 'High': 'C:H'}
        c = cvssCs[getCvssMetricValue(cvssCs, "Confidentiality Impact")]
        cvssIs = {'None': 'I:N', 'Low': 'I:L', 'High': 'I:H'}
        i = cvssIs[getCvssMetricValue(cvssIs, "Integrity Impact")]
        cvssAs = {'None': 'A:N', 'Low': 'A:L', 'High': 'A:H'}
        a = cvssAs[getCvssMetricValue(cvssAs, "Availability Impact")]
        
        colorSubHeading('Temporal Score Metrics')
        cvssEs = {'Not Defined': 'E:X', 
                    'Unproven that exploit exists': 'E:U', 
                    'Proof of concept code': 'E:P', 
                    'Functional exploit exists': 'E:F', 
                    'High': 'E:H'}
        e = cvssEs[getCvssMetricValue(cvssEs, "Exploit Code Maturity")]
        cvssRLs = {'Not Defined': 'RL:X', 
                    'Official fix': 'RL:O', 
                    'Temporary fix': 'RL:T', 
                    'Workaround': 'RL:W', 
                    'Unavailable': 'RL:U'}
        rL = cvssRLs[getCvssMetricValue(cvssRLs, "Remediation Level")]
        cvssRCs = {'Not Defined': 'RC:X', 
                    'Unknown': 'RC:U', 
                    'Reasonable': 'RC:R', 
                    'Confirmed': 'RC:C'}
        rC = cvssRCs[getCvssMetricValue(cvssRCs, "Report Confidence")]
        
        colorSubHeading('Environmental Score Metrics')
        colorSubHeading('Exploit')
        cvssMAVs = {'Not Defined': 'MAV:X', 
                    'Network': 'MAV:N', 
                    'Adjacent Network': 'MAV:A', 
                    'Local': 'MAV:L', 
                    'Physical': 'MAV:P'}
        mAV = cvssMAVs[getCvssMetricValue(cvssMAVs, "Environmental Attack Vector")]
        cvssMACs = {'Not Defined': 'MAC:X', 'Low': 'MAC:L', 'High': 'MAC:H'}
        mAC = cvssMACs[getCvssMetricValue(cvssMACs, "Environmental Attack Complexity")]
        cvssMPRs = {'Not Defined': 'MPR:X', 'None': 'MPR:N', 'Low': 'MPR:L', 'High': 'MPR:H'}
        mPR = cvssMPRs[getCvssMetricValue(cvssMPRs, "Environmental Privileges Required")]
        cvssMUIs = {'Not Defined': 'MUI:X', 'None': 'MUI:N', 'Required': 'MUI:R'}
        mUI = cvssMUIs[getCvssMetricValue(cvssMUIs, "Environmental User Interaction")]
        cvssMSs = {'Not Defined': 'MS:X', 'Unchanged': 'MS:U', 'Changed': 'MS:C'}
        mS = cvssMSs[getCvssMetricValue(cvssMSs, "Environmental Scope")]

        colorSubHeading('Impact')
        cvssMCs = {'Not Defined': 'MC:X', 'None': 'MC:N', 'Low': 'MC:L', 'High': 'MC:H'}
        mC = cvssMCs[getCvssMetricValue(cvssMCs, "Environmental Confidentiality Impact")]
        cvssMIs = {'Not Defined': 'MI:X', 'None': 'MI:N', 'Low': 'MI:L', 'High': 'MI:H'}
        mI = cvssMIs[getCvssMetricValue(cvssMIs, "Environmental Integrity Impact")]
        cvssMAs = {'Not Defined': 'MA:X', 'None': 'MA:N', 'Low': 'MA:L', 'High': 'MA:H'}
        mA = cvssMAs[getCvssMetricValue(cvssMAs, "Environmental Availability Impact")]
        colorSubHeading('Impact Subscore Modifiers')
        cvssCRs = {'Not Defined': 'CR:X', 'Low': 'CR:L', 'Medium': 'CR:M', 'High': 'CR:H'}
        cR = cvssCRs[getCvssMetricValue(cvssCRs, "Environmental Confidentiality Requirement")]
        cvssIRs = {'Not Defined': 'IR:X', 'Low': 'IR:L', 'Medium': 'IR:M', 'High': 'IR:H'}
        iR = cvssIRs[getCvssMetricValue(cvssIRs, "Environmental Integrity Requirement")]
        cvssARs = {'Not Defined': 'AR:X', 'Low': 'AR:L', 'Medium': 'AR:M', 'High': 'AR:H'}
        aR = cvssARs[getCvssMetricValue(cvssARs, "Environmental Availability Requirement")]

        cvssVector = 'CVSS:3.0/'
        cvssVector += aV + '/' + aC + '/' + pR + '/' + uI + '/' + s + '/'
        cvssVector += c + '/' + i + '/' + a + '/'
        cvssVector += e + '/' + rL + '/' + rC + '/'
        cvssVector += mAV + '/' + mAC + '/' + mPR + '/' + mUI + '/' + mS + '/'
        cvssVector += mC + '/' + mI + '/' + mA + '/'
        cvssVector += cR + '/' + iR + '/' + aR
        c = CVSS3(cvssVector)
        returnStr = [str(c.severities()[2]), str(c.scores()[2]), cvssVector]
    return returnStr

def getCvssMetricValue(cvssDict, metricName):
    """Helper to display boilerplate prompts for values"""
    print("What is the " + metricName + "?")
    for (i, opt) in enumerate(list(cvssDict)):
        print("\t" + str(i) + ") " + opt)
    return list(cvssDict)[int(input(" >  "))]

def getMitreAttack():
    """Menu prompting to select MITRE ATT&CK tactic and technique"""
    tactic = ''
    technique = ''
    colorHeader("MITRE ATT&CK")
    path_includes = autorpt_runfrom + '/includes'
    csvFiles = glob(path_includes + '/autorpt-*-attack.csv')
    
    i = 0
    colorNotice('Which MITRE ATT&CK Framework applies?\nPress 99 for main menu.')
    for file in csvFiles:
        colorMenuItem(f'{i}. {file[30:-11]}')
        i = i + 1
    picker = int(input('>  '))
    if 99 == picker:
        mainMenu()
    elif picker > (len(csvFiles)):
        colorNotice('Selection out of range')
        mainMenu()
    else:
        file = csvFiles[picker]
    
    matrix = re.match(r"^autorpt-(\W+)-attack.csv$", str(file))
    df = pd.read_csv(file, index_col=False, engine="python")

    # Get the tactic
    i = 0
    picker = 0
    tactics = df.TACTIC.unique()
    colorNotice('What is the Tactic?\nOr 99 to return to the ATT&CK menu.')
    for tactic in tactics:
        colorMenuItem(f'{i}. {tactic}')
        i = i + 1
    picker = int(input('>  '))
    if 99 == picker:
        getMitreAttack()
    elif picker > len(tactics):
        colorNotice('Selection out of range.')
        mainMenu()
    else:
        tactic = tactics[picker]
    
    # Get the technique
    i = 0
    picker = 0
    techniques = df.query(f'TACTIC == "{tactic}"')[['TECHNIQUE']]
    colorNotice('Pick a Technique?')
    for index, row in techniques.iterrows():
        colorMenuItem(f"{str(i)}.  {str(row.TECHNIQUE)}")
        i = i + 1
    picker = int(input('>  '))
    technique = techniques.iloc[picker, 0]

    # Bug: Upon repeat will not set the correct value for either value.
    # Returns the first selection for both tactic and technique.
    #colorDebug(f"Final values from getMitreAttack() are Tactic {tactic} & Technique: {technique}")
    #colorNotice(f'Is this correct?   Tactic {tactic} & Technique: {technique}')
    #colorMenuItem('1. Yes')
    #colorMenuItem('2. No')
    #picker = int(input('>  '))
    #if 2 == picker:
    #    getMitreAttack()
    
    return [tactic, technique]

def dictToMenu(dictionary):
    """Helper to convert a dictionary to menu item listing"""
    i = 0
    for item in dictionary.split(','):
        colorMenuItem(str(i) + ".  " + item)
        i += 1
    colorMenuItem('99 for main menu')
    return i

def configSectionToMenu(section):
    """Helper to convert a section to menu item listing"""
    i = 0
    items = []
    for item in section:
        colorMenuItem(str(i) + ".  " + section[item])
        items.append(section[item])
        i += 1
    colorMenuItem('99 for main menu')
    return items

def addTarget(ipAddress):
    """Manually get a new IP address for the targets file and copy in a new template"""
    if ipAddress == '':
        # Prompt for target IP address
        colorNotice('Do you know the target IP address?  Or enter "N" to skip.')
        ipAddress = str(input('>  ')).replace(" ", "").lower()
        if ipAddress == 'n':
            colorVerificationFail('No IP Address provided', "An IP Address is required.")
            sys.exit(22)
    # Update targets with a new IP address
    activePath = getActivePath()
    print(f'Injecting {ipAddress} into target file {activePath}/{targetsFile}.')
    with open(f'{activePath}/{targetsFile}', 'a') as f:
        f.write(f'{ipAddress}\n')
    # Copy machine markdown to active report directory
    rptPath = glob(f'{activePath}/report/[0-9]-closing.md')[0]
    filename = rptPath.split('/')[-1]
    filenameBase = filename.split('-')[1]
    filenameOld = f"{activePath}/report/{filename}"
    filenameNew = f"{activePath}/report/"
    filenameNew += f"{str(int(filename.split('-')[0]) + 1)}"
    filenameNew += f"-{filenameBase}"
    try:
        os.rename(filenameOld, filenameNew)
    except:
        colorVerificationFail(f'Failed to rename {filenameOld} to {filenameNew}')
        sys.exit(24)
    # Rename filename to match IP address
    machineFile = f"{str(int(filename.split('-')[0]))}-{ipAddress}.md"
    destMachineMd = f"{activePath}/report/{machineFile}"
    sourceMachineMd = f"{autorpt_runfrom}/templates/training/plain/report/1-renameme.md"
    print(f'Copying machine markdown and renaming to {machineFile}')
    shutil.copyfile(sourceMachineMd, destMachineMd)
    # Log action
    sitrepAuto(f'Added new target: {ipAddress}')

def startup():
    """Initialize a new engagement"""
    colorHeader('Startup')
    # Clear defaults
    targetIp = ''

    # Get Settings.
    # Future: Replace variables throughout with the direct appConfig reference
    studentName = appConfig['Settings']['your_name']
    studentId = appConfig['Settings']['studentid']
    studentEmail = appConfig['Settings']['email']
    style = appConfig['Settings']['style']
    outputFormat = appConfig['Settings']['preferred_output_format']

    # If blank settings prompt for value.  Write out config before proceeding.
    # Prompt to reuse or enter new psuedonym
    if '' == studentName:
        colorNotice(f'What is your name?')
        studentName = (str(input('>  ')))
    if '' == studentEmail:
        colorNotice(f'What is your email?')
        studentEmail = (str(input('>  ')))

    # Write new config.toml
    saveConfig(appConfig)

    # Get the engagement type: training, ctf, exam, bugbounty, pentest
    colorNotice('Select the type of engagement:')
    dictLen = int(dictToMenu(appConfig['Settings']['types']))
    picker = int(input('>  '))
    if 99 == picker:
        mainMenu()
    elif picker >= dictLen:
        mainMenu()
    else:
        engagementType = appConfig['Settings']['types'].split(',')[picker]

    # Set default path for templates. Only exams are unique.
    templates_path = f'{autorpt_runfrom}/templates/training/plain/'
    if 'training' == engagementType:
        # Get the platform
        colorNotice('Select the training platform or 80 to add a custom platform or 99 for main menu')
        
        # List easy to read names
        platformList = []
        for key in appConfig['Training']:
            platformList.append(key)
        
        # List directory friendly names (eg. no spaces)
        providers = configSectionToMenu(appConfig['Training'])
        
        # Get section of the platform
        picker = int(input('>  '))
        if picker == 99:
            mainMenu()
        elif picker == 80:
            colorNotice('Enter the name of the custom platform')
            platform = str(input('>  ')).replace(" ", "").lower()
            platformName = platform
            # Get the box name
            colorNotice('What is the box name?')
            colorNotice('(eg. waldo, kenobi, etc.')
            engagementName = str(input('>  ')).replace(" ", "").lower()
            # Get target IP address
            colorNotice('Do you know the target IP address?  Or enter "N" to skip.')
            targetIp = str(input('>  ')).replace(" ", "").lower()
        elif picker <= 6:    
            # Set the easy to read platform name
            platformName = platformList[picker]

            # Set the directory friendly platform name
            platform = appConfig["Training"][platformName]
            
            # Get the box name
            colorNotice('What is the box name?')
            colorNotice('(eg. waldo, kenobi, etc.')
            engagementName = str(input('>  ')).replace(" ", "").lower()
            # Get target IP address
            colorNotice('Do you know the target IP address?  Or enter "N" to skip.')
            targetIp = str(input('>  ')).replace(" ", "").lower()
        else:
            # Set the easy to read platform name
            platformName = platformList[picker]

            # Set the directory friendly platform name
            #platform = appConfig["Training"][platformName]
            platform = 'offensivesecurity'
            engagementName = providers[picker]
            templates_path = f"{autorpt_runfrom}/templates/training/{engagementName}"
    elif 'bugbounty' == engagementType:
        templates_path = f'{autorpt_runfrom}/templates/training/bugbounty/'
        # Get the platform
        colorNotice('Enter the platform or company name:')
        providers = configSectionToMenu(appConfig['Bug Bounty'])
        picker = int(input('>  '))
        if 3 == picker: 
            platformName = str(input('What is your penetration testing company name?  '))
        else:
            platformName = providers[picker]
        platform = platformName.lower()
        platform = platform.replace("'", "")
        platform = platform.replace('"', "")
        platform = platform.replace('`', '')
        platform = platform.replace('/', '')
        platform = platform.replace('\\', '')
        platform = platform.replace(" ", "")
        # Get the program name
        colorNotice('What is the program name?')
        colorNotice('(eg. Tesla, Domain.com, etc.')
        engagementName = str(input('>  ')).replace('\s+', '').lower()
        engagementName = engagementName.replace("'", "")
        engagementName = engagementName.replace('"', "")
        engagementName = engagementName.replace('`', '')
        engagementName = engagementName.replace(' ', '')
    elif 'ctf' == engagementType:
        # Get the engagement name
        colorNotice('What is the name of this CTF event?')
        platform = str(input('>  '))
        platformName = platform
        platform = platform.lower()
        platform = platform.replace("'", "")
        platform = platform.replace('"', "")
        platform = platform.replace('`', '')
        platform = platform.replace('/', '')
        platform = platform.replace('\\', '')
        platform = platform.replace(" ", "")
        # Get the engagement name
        colorNotice('What is the team name?')
        engagementName = str(input('>  ')).lower()
        engagementName = engagementName.replace("'", "")
        engagementName = engagementName.replace('"', "")
        engagementName = engagementName.replace('`', '')
        engagementName = engagementName.replace(' ', '')
        templates_path = f'{autorpt_runfrom}/templates/training/plain/'
    elif 'exam' == engagementType:
        # pick the exam
        colorNotice('Select the exam')
        i = 0
        exams = []
        for item in appConfig['Exams']:
            examName = appConfig['Exams'][item].split(',')[1]
            colorMenuItem(str(i) + ".  " + examName)
            exams.append(item)
            i += 1
        colorMenuItem('99. for main menu')
        picker = int(input('>  '))
        if 99 == picker:
            mainMenu()
        platform = appConfig['Exams'][exams[picker]].split(',')[0]
        platformName = appConfig['Exams'][exams[picker]].split(',')[1]
        engagementName = exams[picker]
        templates_path = f'{autorpt_runfrom}/templates/{engagementName}/'
        # Prompt for student ID, email, name, etc.
        if '' == studentId:
            colorNotice(f'What is your student ID?')
            studentId = (str(input('>  ')))
    elif 'pentest' == engagementType:
        # Least tested option
        # Company performing the test
        platformName = str(input('What is your penetration testing company name?  '))
        platform = platformName.lower()
        platform = platform.replace("'", "")
        platform = platform.replace('"', "")
        platform = platform.replace('`', '')
        platform = platform.replace('/', '')
        platform = platform.replace('\\', '')
        platform = platform.replace(" ", "")
        # Client name
        colorNotice('What is the client name?')
        engagementName = str(input('>  ')).replace('\s+', '').lower()
        engagementName = engagementName.replace("'", "")
        engagementName = engagementName.replace('"', "")
        engagementName = engagementName.replace('`', '')
        engagementName = engagementName.replace(' ', '')
    else:
        mainMenu()
    
    # Set the preferred output format or set to null
    if '' == outputFormat:
        engagementFormat = ''
        colorNotice('Do you have a preferred output format for the final report?  Enter for none.')
        dictLen = int(dictToMenu(appConfig['Settings']['output_formats']))
        picker = int(input('>  '))
        if 99 == picker:
            mainMenu()
        elif picker >= dictLen:
            mainMenu()
        else:
            outputFormat = appConfig['Settings']['output_formats'].split(',')[picker]

    # Set timestamp for this engagement for uniqueness
    timestamp = datetime.datetime.now().strftime('%Y%m%d')
    
    # Compile the engagement string and directory path to create
    thisEngagement = f'{engagementType}'
    thisEngagement += f'-{platform}'
    thisEngagement += f'-{engagementName}'
    thisEngagement += f'-{timestamp}'

    thisDir = appConfig['Paths']['pathwork']
    thisDir += f'/{engagementType}'
    thisDir += f'/{platform}'
    thisDir += f'/{engagementName}'
    thisDir += f'-{timestamp}'

    # Copy the template to the engagement directory
    try:
        err = shutil.copytree(templates_path, f'{thisDir}/')
    except OSError as exc: # python >2.5
        colorVerificationFail("[!]", f"Copytree templates failed. {err}")
        try:
            err = shutil.copy(templates_path, f'{thisDir}/')
        except OSError as exc:
            colorVerificationFail("[!]", f"Copy templates failed. {err}")
            sys.exit(5)
    
    if "training" == engagementType and re.search('plain', templates_path, flags=0):
        os.rename(f'{thisDir}/report/1-renameme.md', f'{thisDir}/report/1-{engagementName}.md')
        if len(targetIp) >= 7:
            with open(f'{thisDir}/{targetsFile}', 'w') as t:
                t.write(targetIp + '\n')
    
    # Update sessions file
    # Set active
    session['Current']['active'] = thisEngagement
    
    # Set engagement settings
    session[thisEngagement] = {}
    session[thisEngagement]['path'] = thisDir
    session[thisEngagement]['platform'] = platformName
    session[thisEngagement]['type'] = engagementType
    session[thisEngagement]['student_id'] = studentId
    session[thisEngagement]['student_name'] = studentName
    session[thisEngagement]['student_email'] = studentEmail
    session[thisEngagement]['style'] = style
    session[thisEngagement]['engagement_name'] = engagementName
    session[thisEngagement]['output_format'] = outputFormat
    session[thisEngagement]['status'] = 'Started'
    session[thisEngagement]['start'] = str(datetime.datetime.now())
    session[thisEngagement]['end'] = ''
    #colorDebug(f"Session: {session}")
    saveEnagements()

    # Journal entry in sitrep
    msg = f'Startup initiated for {engagementType} as {engagementName}'
    sitrepAuto(msg)
    msg = f'New working directory is {thisDir}'
    sitrepAuto(msg)

    # Display directory tree
    if 'exam' == engagementType:
        colorNotice("You will need to manually update the targets file.")
    colorNotice("Templates successfully copied to report directory.  Here's the new structure:")
    colorNotice(thisDir)
    paths = DisplayablePath.make_tree(Path(thisDir))
    for path in paths:
        print(path.displayable())
    time.sleep(0.5)

def finalize():
    """Create the final report by combining all numbered markdown files and calling pandoc"""
    toArchive = 'No'
    active = session['Current']['active']
    engagementType = session[active]["type"]
    email = session[active]['student_email']
    author = session[active]['student_name']
    student_id = session[active]['student_id']
    style_name = appConfig['Settings']['style']
    rptFormat = session[active]['output_format']
    activePath = session[active]["path"]
    rpt_base = f"{activePath}/report/"
    rpt_date = datetime.datetime.now().strftime('%Y-%m-%d')
    targetName = active.split('-')[0]
    platformName = rpt_base.split('/')[-4]
    
    # Change to working directory
    os.chdir(rpt_base)

    # Ensure the latest ports file exists
    portsFile = f"{rpt_base}{portsSpreadsheet}"
    if not os.path.exists(portsFile):
        ports()
    # Read in the ports spreadsheet
    ports_table = pd.read_excel(portsFile, sheet_name='All Ports').to_markdown()

    # FUTURE FEATURE IS TO AUTOMATICALLY ADD THE VULNS TABLE'
    portsFile = f"{rpt_base}{vulnsCsv}"
    if os.path.isfile(portsFile):
        fields = ['CvssSeverity','IpAddress','Port','Name','Remediation']
        vulns_table = pd.read_csv(portsFile, 
                                  usecols=fields, 
                                  sep=",", 
                                  engine="python").to_markdown()
        vulns_all = pd.read_csv(portsFile, sep=",", engine="python")
    else:
        vulns_table = 'No vulnerabilities were discovered.'

    # Student info only applies for some exams (eg OffSec)
    if 'training' == engagementType:
        rpt_name = "training_" + targetName + "_Report"
    else:
        rpt_name = f"{targetName.upper()}_" + student_id + "_Exam_Report"

    if 'exam' == engagementType:
        if student_id == '':
            colorNotice("\nWhat is your student ID, if required?\n(eg. OS-12345, N/A)")
            student_id = str(input('>  '))
        else:
            colorNotice("Student ID pulled from config file as " + student_id)
    
    if email == '':
        colorNotice("What is your full email address?")
        email = str(input('>  '))
    else:
        colorNotice("Email address pulled from config file as " + email)

    if author == '':
        colorNotice("What is your name?")
        author = str(input('>  '))
    else:
        colorNotice("Author pulled from config file as " + author)
    
    # Rename as rptMarkdownFile
    rpt_filename = rpt_base + rpt_name + ".md"
    
    # Set output file format
    if rptFormat == '':
        i = 0
        print("From these options, Pick an output format:")
        for ext in supported_filetypes.split(','):
            colorMenuItem(f"{str(i)} ) {ext}")
            i += 1
        picked = int(input('>  '))
        rptFormat = supported_filetypes[picked].lower()

    if rptFormat in ["commonmark_x", "jira", "gfm"]:
        rpt_extension = 'md'
    else:
        rpt_extension = rptFormat
    # Is archive needed?
    if '7z' in rpt_extension:
        toArchive = 'yes'
        rpt_extension = rpt_extension.split("+")[0]
        rptFullPath = rpt_base + rpt_name + "." + rpt_extension
    else:
        rptFullPath = rpt_base + rpt_name + "." + rpt_extension
    
    # FUTURE: Merge and build lab report
    # The report title, other metadata, and closing require changes
    # Since there are two directories needing reports loop over each or function call with an object.
    # if type == "exam":
    #     for dir in lab_dir, exam_dir:
    #         cd dir # hopefully this will pick up relative image locations

    # Create exam report
    msg = f'Creating final report.  toArchive? {toArchive}  Ext: {str(rpt_extension)}  File: {rptFullPath}'
    sitrepAuto(msg)
    
    # The merged, unified markdown file is not a primary source.
    # Remove of it already exists. Previous failed attempt.
    if os.path.exists(rpt_filename):
        try:
            os.remove(rpt_filename)
        except:
            colorVerificationFail("Error", "removing existing report file: " + rpt_filename)
            sys.exit(6)

    # Merge markdown sections into a single mardown for pandoc later
    md_file_list = glob(rpt_base + '[0-9]*.md')
    md_file_list = sorted(md_file_list)
    for file in md_file_list:
        with open(file, 'r+') as f:
            file_contents = f.read()
            file_contents = re.sub('BOILERPLATE_AUTHOR', author, file_contents)
            file_contents = re.sub('BOILERPLATE_EMAIL', email, file_contents)
            file_contents = re.sub('BOILERPLATE_DATE', rpt_date, file_contents)
            file_contents = re.sub('BOILERPLATE_PORTS', ports_table, file_contents)
            file_contents = re.sub('BOILERPLATE_VULNS', vulns_table, file_contents)
            if "training" == engagementType:
                # Future feature: auto add keywords from directory: hackthebox,popcorn,write-up,training
                # get platform
                file_contents = re.sub('BOILERPLATE_PLATFORM', platformName, file_contents)
                file_contents = re.sub('BOILERPLATE_TARGET', targetName, file_contents)
                file_contents = re.sub('BOILERPLATE_HOSTNAME', targetName, file_contents)
                file_contents = re.sub('BOILERPLATE_OSID', '', file_contents)
            else:
                file_contents = re.sub('BOILERPLATE_OSID', student_id, file_contents)
                
            # Some markdown for pasted images are incompatable with Pandoc 
            # and results in no images reaching the final report.
            # Rewrite all pasted screenshots from this "![[Pasted_image_A.png]]" 
            # to this format "![Pasted_image_A.png](Pasted_image_A.png)"
            file_contents = re.sub(r'\!\[\[Pasted_image_(\w+).png\]\]', 
                                   r'![Pasted_image_\1.png](Pasted_image_\1.png)', 
                                   file_contents)
            # Write modifed contents with boilerplate value replacements
            with open(rpt_filename, 'a') as result:
                result.write(file_contents + '\n')

    if style_name == '':
        style_name = getPandocStyle()
    else:
        colorNotice("Code block style pulled from config file as " + style_name)
    
    colorNotice(f"Generating report {rptFullPath}")
    # Build the Pandoc command for generating the report
    extentionsWithoutTemplate = []
    extentionsWithoutTemplate = appConfig['Settings']['no_template'].split(',')

    cmd = '/usr/bin/pandoc ' + rpt_filename
    cmd += ' --output=' + rptFullPath
    cmd += ' --from markdown+yaml_metadata_block+raw_html'
    if not rptFormat in extentionsWithoutTemplate:
        cmd += ' --template' + ' eisvogel'
    cmd += ' --table-of-contents' 
    cmd += ' --toc-depth' + ' 6'
    cmd += ' --top-level-division=chapter'
    cmd += ' --number-sections'
    cmd += ' --wrap=auto'
    cmd += ' --highlight-style ' + style_name
    # Helpful for debugging the pandoc command: 
    #colorDebug(f"cmd:\n{cmd}")

    try:
        p = subprocess.getoutput(cmd)
    except:
        colorFail("[!]", f"Failed to generate PDF using pandoc.\n{p}")
        sys.exit(10)
    colorNotice(p)

    if 'yes' == toArchive:
        archive_file = rpt_base + rpt_name + ".7z"
        colorNotice("Generating 7z archive " + archive_file)
        cmd = '/usr/bin/7z a ' + archive_file + ' ' + rptFullPath
        try:
            p = subprocess.getoutput(cmd)
        except:
            colorFail("[!]", f"Failed to generate 7z archive\n{p}")
            sys.exit(15)
        # Debug 7zip output: colorNotice(p)
    # Log the action taken
    msg = f"Report finalized as {rptFullPath}"
    sitrepAuto(msg)

    # Update the engagement status
    active = session['Current']['active']
    session[active]['end'] = str(datetime.datetime.now())
    session[active]['status'] = 'Finalized'
    saveEnagements()

# Obsolete
def getActivePath():
    """Deprecated means of getting the active engagement path"""
    active = session['Current']['active']
    if 'None' == active:
        colorNotice('No active engagement exists.  Use startup to create a new engagement.')
        sys.exit(30)
    else:
        return session[active]["path"]
# Obsolete
def getActiveAll():
    """Deprecated means of getting the active engagement name"""
    active = session['Current']['active']
    if 'None' == active:
        colorNotice('No active engagement exists.  Use startup to create a new engagement.')
        sys.exit(30)
    else:
        return f"{session['Engagements'][active]}"

def getPandocStyle():
    """Selector of the code syntax highlight style"""
    colorNotice("\nFrom the following list, pick a syntax highlight style for code blocks?")
    colorNotice("   Recommendation: lighter styles are easier to read and use less ink if printed.")
    colorNotice("   Dark styles include: espresso, zenburn, and breezedark.")
    colorNotice("   Light styles include: pygments, tango, kate, monochrome, haddock")
    i = 0
    style_list = {}
    p = str(subprocess.run(["pandoc", "--list-highlight-styles"], 
                            check=True, 
                            universal_newlines=True, 
                            capture_output=True).stdout)
    output = p.splitlines(False)
    for s in output:
        colorList('\t' + str(i) + ". " + s)
        style_list[i] = s
        i += 1
    style_id = int(input('>  '))
    if style_id > i or style_id < 0:
        colorNotice('Invalid selection')
        getPandocStyle()
    return style_list[style_id]

def getNmapFile(target):
    """A listing of known good nmap output files. In order: AutoRecon, nmapAutomator, and Reconnoitre."""
    nmapFiles = [f"_full_tcp_nmap.txt",
                 f"_quick_tcp_nmap.txt",
                 f"Full_{target}.nmap",
                 f"{target}.quick.nmap",
                 f"{target}.nmap"]
    
    for name in nmapFiles:
        for root, dirs, files in os.walk(getActivePath()):
            if name in files:
                nmapFile = os.path.join(root, name)
                return nmapFile

def ports():
    """Display the ports, replace ports spreadsheet."""
    portsFile = f"{getActivePath()}/report/{portsSpreadsheet}"
    if os.path.isfile(portsFile):
        os.remove(portsFile)
    # Look for nmap output files associated with each target IP address
    with open(f"{getActivePath()}/{targetsFile}", 'r', encoding='utf-8', newline='') as t:
        allPorts = pd.DataFrame({})
        targets = t.readlines()
        for target in targets:
            target = target.strip()
            nmapFile = getNmapFile(target)
            if nmapFile == None:
                print('[!] Unable to find any nmap files.')
                break

            df = pd.DataFrame({})
            ip = ''
            port = '' 
            state = ''
            service = '' 
            version = ''
            with open(nmapFile, 'r', encoding='utf-8', newline='') as n:
                nmapContents = n.readlines()
                n.close()
            for line in nmapContents:
                if re.match(r"^Nmap scan report for ", line):
                    ip = line.strip().replace('Nmap scan report for ', '')
                elif re.match(r"^\d+.*$", line):
                    fields = line.strip().split()
                    # 0:port, 1:state, 2:service, 3:reason(skip), 4:version(glob)
                    port = fields[0]
                    state = fields[1]
                    service = fields[2]
                    #version = ' '.join(fields[4:])
                    version = re.sub(r'\(.*\)', '', ' '.join(fields[4:]))

                    if service == 'unrecognized':
                        continue
                    
                    newRow = {'IPADDRESS': ip,
                            'PORT': port, 
                            'STATE': state, 
                            'SERVICE': service, 
                            'VERSION': version}
                    df = df.append(newRow, ignore_index = True)
                    allPorts = allPorts.append(newRow, ignore_index = True)
            # Create worksheet per target
            colorVerification(target, f'Port Count: {str(len(df.index))}')
            #colorList(df.to_markdown())

            with pd.ExcelWriter(portsFile, engine='openpyxl') as writer:
                if os.path.exists(portsFile):
                    book = openpyxl.load_workbook(portsFile)
                else:
                    book = openpyxl.Workbook()
                
                writer.book = book
                sheetActive = book.active
                if 'Sheet' in book.sheetnames:
                    del book['Sheet']

                try:
                    df.to_excel(writer, sheet_name=target, index=False)
                except:
                    colorFail("[e]", "Unable to write to xlsx file.")
                try:
                    writer.save()
                except:
                    colorFail("[e]", "Unable to save xlsx file.")
                try:
                    writer.close()
                except:
                    colorFail("[e]", "Unable to close xlsx file.")
        colorList(allPorts.to_markdown())

        with pd.ExcelWriter(portsFile, engine='openpyxl') as writer:
            if os.path.exists(portsFile):
                book = openpyxl.load_workbook(portsFile)
            else:
                book = openpyxl.Workbook()
            
            writer.book = book
            sheetActive = book.active
            if 'Sheet' in book.sheetnames:
                del book['Sheet']

            try:
                allPorts.to_excel(writer, sheet_name='All Ports', index=False)
            except:
                colorFail("[e]", "Unable to write to xlsx file.")
            try:
                writer.save()
            except:
                colorFail("[e]", "Unable to save xlsx file.")
            try:
                writer.close()
            except:
                colorFail("[e]", "Unable to close xlsx file.")
    
    # Update the engagement status
    active = session['Current']['active']
    session[active]['status'] = 'In-process'
    saveEnagements()
    return allPorts

def sitrepAuto(msg):
    """Automatically, without prompting, write the msg to the sitrep file."""
    d = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    sitrepFile = f"{getActivePath()}/report/{sitrepLog}"
    if os.path.exists(sitrepFile):
        with open(sitrepFile, 'a', encoding='utf-8', newline='') as f:
            f.write(f'{d} - {msg}\n')
            f.close()
    else:
        with open(sitrepFile, 'w', encoding='utf-8', newline='') as f:
            f.write(f'{d} - {msg}\n')
            f.close()
    #colorNotice('sitrep logged')

def sitrepList():
    """Display the contents of the sitrep file"""
    sitrepFile = f"{getActivePath()}/report/{sitrepLog}"
    if os.path.isfile(sitrepFile):
        colorHeader("SITREP Log Entries")
        with open(sitrepFile) as f: 
            s = f.readlines()
            f.close()
        for l in s:
            fields = l.strip().split(" - ")
            colorVerification(fields[0], fields[1])
    else:
        print(f'{term.white}Sitrep file is empty.{term.normal}\n\n')

def sitrepNew():
    """Manually prompt for the sitrep message."""
    msg = str(input('What is your current status? '))
    sitrepAuto(msg)
    sitrepMenu()

def sitrepMenu():
    """A stream of status journal."""
    colorHeader('SITREP  (Situation Report)')
    colorMenuItem('1. List all sitrep entries')
    colorMenuItem('2. Add new sitrep log entry\n')
    colorMenuItem('3. Main Menu')
    colorMenuItem('4. Quit')
    sitrepAction = int(input('>  '))
    if 4 == sitrepAction:
        sys.exit(0)
    elif 3 == sitrepAction:
        mainMenu()
    elif 1 == sitrepAction:
        sitrepList()        
    elif 2 == sitrepAction:
        sitrepNew()
    sitrepMenu()

def vuln():
    """Menu for vulnerabilities"""
    colorHeader('Vulnerabilities')
    colorMenuItem("1. Add a new vulnerability")
    colorMenuItem("2. List all vulnerabilities")
    colorMenuItem("3. Modify an existing vulnerability")
    colorMenuItem("4. Remove a vulnerability\n")
    colorMenuItem('5. Main Menu')
    colorMenuItem("6. quit")
    vuln_selection = int(input("> "))
    if 1 == vuln_selection:
        vulnAdd()
    elif 2 == vuln_selection:
        vulnList()
    elif 3 == vuln_selection:
        vulnModify()
    elif 4 == vuln_selection:
        vulnRemove()
    elif 5 == vuln_selection:
        mainMenu()
    elif 6 == vuln_selection:
        sys.exit(0)
    else:
        vuln()

def vulnAdd():
    """Manually prompt for details of a validated vulnerability."""
    colorHeader("Add Vulnerability")
    i = 0
    target = ""
    targetFile = f"{getActivePath()}/{targetsFile}"
    colorDebug(f'Target File: {targetFile}')
    if os.path.isfile(targetFile):
        colorNotice("For which target?\nOr '99' to go back to the menu.")
        with open(targetFile) as f: 
            targets = f.readlines()
            f.close()
        for target in targets:
            target = target.strip()
            #print(f'  {i}.  {term.yellow}{target}{term.normal}')
            colorMenuItem(f"{i}.  {target}")
            i = i + 1
        targetId = int(input(">  "))
        if 99 == targetId:
            vuln()
        else:
            target = targets[targetId].strip()
    else:
        print(f'targets file is empty. Please add IP addresses to the targets file.\n\n')
        vuln()
    
    colorNotice("What is the port number [0-65535]?")
    port = int(input('>  '))
    if port < 1 or port > 65535:
        colorNotice("Number is outside the range of acceptable port numbers: 0 - 65535.")
        vuln()
    
    colorNotice("What is the name for this vulnerability?\n(eg. Remote code injection in Vendor_Product_Component)")
    vulnName = str(input('>  '))
    
    colorNotice("Describe the business impact: ")
    vulnImpact = str(input('>  '))
    
    colorNotice("What is the remediation?")
    remediation = str(input('>  '))
    
    colorNotice("Do you have a comment for where you left off? ")
    vulnComment = str(input('>  '))
    if len(vulnComment) > 0:
        sitrepAuto(vulnComment)
    
    rawCvss = getCvss3Score()
    cvssSeverity = rawCvss[0]
    cvssScore = rawCvss[1]
    cvssVector = rawCvss[2]
    mitreAttack = getMitreAttack()
    vulnMitreTactic = mitreAttack[0]
    vulnMitreTechnique = mitreAttack[1]
    
    colorNotice("\n-----------------------------\n  Verify the data entered.\n-----------------------------")
    colorVerification('[Target]                ',target)
    colorVerification('[Port]                  ',port)
    colorVerification('[Name]                  ',vulnName)
    colorVerification("[CVSS Overall Score]    ", str(cvssScore))
    colorVerification("[CVSS Severity]         ", cvssSeverity)
    colorVerification("[Business Impact]       ", vulnImpact)
    colorVerification("[Remediation    ]       ", remediation)
    colorVerification("[Comment]               ", vulnComment)
    colorVerification('[MITRE ATT&CK Tactic]   ', vulnMitreTactic)
    colorVerification('[MITRE ATT&CK Technique]', vulnMitreTechnique)
    
    checkPoint = str(input("\nAre these values correct? [Y|N]  > ")).upper()
    if checkPoint == "Y":
        row = f'{target},{port},{vulnName},{vulnImpact},{remediation},{vulnComment},{cvssScore},{cvssSeverity},{cvssVector},{vulnMitreTactic},{vulnMitreTechnique}'
        vulnCsvNewRow(row)
    else:
        print("[!] Reseting values")
        vulnName = ''
        vulnImpact = ''
        remediation = ''
        vulnCvss = ''
        vulnMitreTactic = ''
        vulnMitreTechnique = ''
        vulnComment = ''
        vulnAdd()
    # Return to the vulnerability menu
    vuln()

def vulnCsvNewRow(row):
    """Formats the vulnerability row and stores in the spreadsheet."""
    vulnsFile =  f"{getActivePath()}/report/{vulnsCsv}"
    if not os.path.isfile(vulnsFile):
        headings = 'IpAddress,Port,'
        headings += 'Name,Impact,Remediation,Comment'
        headings += ',CvssScore,CvssSeverity,CvssVector'
        headings += ',MitreTactic,MitreTechnique'
        with open(vulnsFile, 'a', encoding='utf-8') as f:
            f.write(headings + "\n")
            f.write(row + "\n")
            f.close()
    else:
        with open(vulnsFile, 'a', encoding='utf-8') as f:
            f.write(row + "\n")
            f.close()
    msg = f'Added new vulnerability: {str(row)}'
    sitrepAuto(msg)
    # Update the engagement status
    active = session['Current']['active']
    session[active]['status'] = 'In-process'
    saveEnagements()

def vulnList():
    """Displays a list of current vulnerabilities from the spreadsheet."""
    colorHeader("List of Current Vulnerabilities")
    vulnsFile =  f"{getActivePath()}/report/{vulnsCsv}"
    if os.path.exists(vulnsFile):
        df = pd.read_csv(vulnsFile, sep=",", engine="python") # , index_col=False
        colorList(df.to_markdown())
    else:
        print("0 vulnerabilities")
    if len(sys.argv) == 3 and 'list' == sys.argv[2]:
        sys.exit(0)    
    else:
        vuln()

def vulnModify():
    """Modify a vulnerability"""
    print("\n")
    vulnsFile =  f"{getActivePath()}/report/{vulnsCsv}"
    try:
        vm = pd.read_csv(vulnsFile)
    except:
        colorNotice('No vulnerabilities logged to modify.')
        vuln()
    rowCount = len(vm.index)
    headings = list(vm.columns.values)
    colorList(vm.to_markdown())
    vulnId = int(input("\nPick an entry to modify or '99' to go back to the menu:  "))
    if 99 == vulnId:
        vuln()
    print("\n" + str(headings))
    fieldId = str(input("Type a column name to modify or '99' to go back to the menu:  "))
    if '99' == fieldId:
        vuln()
    if fieldId not in headings:
        print("INVALID HEADING")
        vuln()
    if 'CVSS' == fieldId:
        newValue = float(input("What is the new value?  "))
    else:
        newValue = str(input(f'What is the new value?  '))
    vm.at[vulnId, fieldId] = newValue
    print(vm)
    msg = f"Modified vulnerability {fieldId} to: {newValue} for {str(vm.at[vulnId, 'Name'])}"
    colorNotice(msg)
    sitrepAuto(msg)
    with open(vulnsFile, 'w', newline='') as f:
        vm.to_csv(f, index=False)
        f.close()
    time.sleep(2)
    vuln()

def vulnRemove():
    """Remove a stored vulnerability"""
    # TBD: Replace with modification of vulnModify
    i = 0
    print("\n")
    vulnsFile =  f"{getActivePath()}/report/{vulnsCsv}"
    try:
        r = csv.reader(open(vulnsFile))
    except:
        colorNotice('No vulnerabilities logged to remove.')
        vuln()
    rows = list(r)
    for row in rows:
        colorList(str(i) + ") " + str(row))
        i += 1
    vulnId = int(input("\nPick an entry to remove or '99' to go back to the menu:  "))
    if 99 == vulnId:
        vuln()    
    # Bug: Get vuln name and log it with sitrep
    msg = "Remove vulnerability: " + str(rows[vulnId])
    sitrepAuto(msg)
    del rows[vulnId]
    with open(vulnsFile, 'w', newline='') as f:
        writer = csv.writer(f, delimiter=',')
        writer.writerows(rows)
        f.close()
    vuln()

def listEngagements():
    """List all engagement names and paths"""
    for e in session.sections():
        if e not in ['DEFAULT', 'Current']:
            colorVerification(e, session[e]['path'])
    settingsMenu()

def settingsMenu():
    """Submenu for settings"""
    colorHeader('Settings')
    colorMenuItem('1. Application-level settings')
    colorMenuItem('2. Engagement settings')
    colorMenuItem('3. Back to main menu')
    colorMenuItem('4. Quit')
    picker = int(input('>  '))
    
    if 1 == picker:
        # Application-level settings
        picker = 0
        colorSubHeading('Current Settings')
        colorMenuItem(f"1) Engagements will be stored in {str(appConfig['Paths']['pathwork'])}")
        colorMenuItem(f"2) Your name: {str(appConfig['Settings']['your_name'])}")
        colorMenuItem(f"3) Your student ID: {str(appConfig['Settings']['studentid'])}")
        colorMenuItem(f"4) Your email address: {str(appConfig['Settings']['email'])}")
        colorMenuItem(f"5) Preferred report format: {str(appConfig['Settings']['preferred_output_format'])}")
        colorMenuItem(f"6) Code block style: {str(appConfig['Settings']['style'])}")
        colorMenuItem(f"7) Settings menu")
        colorMenuItem(f"8) Main menu")
        colorMenuItem(f"\nPick a number to modify its setting")
        picker = int(input('>  '))

        if 8 <= picker:
            # return to main menu
            mainMenu()
        elif 1 == picker:
            # Set new engagements working directory
            colorNotice('What is the path to store future engagement subdirectories?')
            picker = str(input('>  '))
            if not os.path.isdir(picker):
                colorVerificationFail('[!]', f'{picker} is not a valid directory.  Creating...')
                
                # If umask is not set, incorrect permissions will be assigned on mkdir
                os.umask(0o007)
                try:
                    os.mkdir(picker, 0o770)
                except:
                    colorFail('[e]', f'Unable to create directory: {picker} ')
                    sys.exit(20)
            appConfig['Paths']['pathwork'] = picker
            saveConfig(appConfig)
        elif 2 == picker:
            # Set author name
            colorMenuItem('What is your full name as the report author?')
            picker = str(input('>  '))
            appConfig['Settings']['your_name'] = picker
            saveConfig(appConfig)
        elif 3 == picker:
            # Set student ID
            colorMenuItem('What is your student ID?')
            picker = str(input('>  '))
            appConfig['Settings']['studentid'] = picker
            saveConfig(appConfig)
        elif 4 == picker:
            # Set email address
            colorMenuItem('What is your email address?')
            picker = str(input('>  '))
            if (re.match(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', picker)):
                appConfig['Settings']['email'] = picker
                saveConfig(appConfig)
            else:
                colorNotice('Not an email formatted string.  Try again.')
                settingsMenu()
        elif 5 == picker:
            # Set preferred report format
            colorMenuItem(f"What is your preferred report format?  {appConfig['Settings']['preferred_output_format']}")
            i = 0
            for filetype in supported_filetypes.split(','):
                colorMenuItem(f'{i}) {filetype}')
                i += 1
            picker = int(input('>  '))
            if picker <= i:
                appConfig['Settings']['preferred_output_format'] = supported_filetypes.split(',')[picker]
                saveConfig(appConfig)
            else:
                colorNotice('Invalid option.')
                settingsMenu()
        elif 6 == picker:
            # Set code block style
            appConfig['Settings']['style'] = getPandocStyle()
            saveConfig(appConfig)
        elif 7 == picker:
            # return to settings menu
            settingsMenu()
        else:
            # Unknown selection, return to settings menu
            settingsMenu()
    elif 2 == picker:
        # Engagement settings
        colorSubHeading('Engagement Settings')
        colorNotice(f"Active engagement: {session['Current']['active']}")
        #colorNotice(f"Total engagement: {str(len(session['Engagements']))}\n")
        colorMenuItem('1. Set a new active engagement')
        colorMenuItem('2. List all engagements')
        colorMenuItem('3. Back to main menu')
        colorMenuItem('4. Quit')
        picker = int(input('>  '))
        if picker == 3:
            mainMenu()
        elif picker == 4:
            sys.exit(0)
        elif picker > 4:
            settingsMenu()
        elif picker == 1:
            i = 0
            engagements = {}
            colorNotice('Pick a new active engagement')
            for e in session.sections():
                if e not in ['DEFAULT', 'Current']:
                    # Future Feature: Would be nice to exclude finalized items
                    colorMenuItem(f"{i}) {e}")
                    engagements[i] = e
                    i += 1
            picker = int(input('>  '))
            if picker > i:
                settingsMenu()
            else:
                session['Current']['active'] = engagements[picker]
                saveEnagements()
                settingsMenu()
        elif picker == 2:
            listEngagements()
        sys.exit(255)

        colorNotice('Pick an engagement to make it active:')
        i = 0
        for s in session['Engagements']:
            colorMenuItem(f"{i}) {s} - {session['Engagements'][s].split(',')[0]}")
            i += 1
        picker = int(input('>  '))
        if picker <= i:
            colorNotice('set active, save')
            saveEnagements()
        else:
            colorVerificationFail('[!]', 'Invalid selection.')
            settingsMenu()
    elif 4 == picker:
        # quit
        sys.exit(23)
    else:
        # return to main menu
        mainMenu()
    settingsMenu()

def mainMenu():
    """Primary menu"""
    clearScreen()
    banner()
    colorHeader('Main Menu')
    colorMenuItem('1. Startup')
    colorMenuItem('2. Vulnerabilities')
    colorMenuItem('3. Ports')
    colorMenuItem('4. SitRep Log')
    colorMenuItem('5. Finalize')
    colorMenuItem('6. Settings')
    colorMenuItem('7. Quit')
    picker = int(input('>  '))

    if 1 == picker:
        startup()
    elif 2 == picker:
        vuln()
    elif 3 == picker:
        ports()
    elif 4 == picker:
        sitrepMenu()
    elif 5 == picker:
        finalize()
    elif 6 == picker:
        settingsMenu()
    elif 7 == picker:
        sys.exit(0)
    else:
        mainMenu()

def params(argv):
    """Set routing action based on argument.  Otherwise, display help."""
    action = sys.argv[1]
    if action == '-a' or action == '--add' or action == 'add':
        # Add a new target host to the engagement
        if len(sys.argv) == 3:
            msg = sys.argv[2]
        else:
            msg = ''
        addTarget(msg)
    elif action == '-h' or action == '--help' or action == 'help':
        helper()
    elif action == '-s' or action == 'startup' or action == '--startup':
        startup()
    elif action == '-f' or action == 'finalize' or action == '--finalize':
        finalize()
    elif action == 'whathaveidone' or action == 'stats':
        whathaveidone()
    elif action == '-v' or action == 'vuln' or action == '--vuln':
        # Record a confirmed vulnerability
        if len(sys.argv) == 3 and 'list' == sys.argv[2]:
            vulnList()
        else:
            vuln()
    elif action == '-i' or action == 'sitrep' or action == '--sitrep':
        # Situation report actions
        if len(sys.argv) == 3 and 'list' == sys.argv[2]:
            sitrepList()
        elif len(sys.argv) > 3:
            msg = " ".join(sys.argv[2:])
            sitrepAuto(msg)
        else:
            sitrepMenu()
    elif action == '-p' or action == 'ports' or action == '--ports':
        # Update ports and service versions
        ports()
    elif action == 'active':
        if 'None' == session['Current']['active']:
            colorNotice(f"No engagements.  Run 'autorpt.py startup' to create an engagement.")
        else:
            colorNotice(f"The active engagement is: {session['Current']['active']}")
            colorNotice(session[session['Current']['active']]['path'])
    else:
        mainMenu()

def loadAppConfig(pathConfig, appConfigFile):
    """Read application-level settings configuration file"""
    # Exit without configuration file
    if not os.path.isdir(pathConfig):
        # If umask is not set, incorrect permissions will be assigned on mkdir
        os.umask(0)
        try:
            os.mkdir(pathConfig, 0o770)
        except:
            colorFail('[e]', f'Unable to create directory for AutoRpt settings: {pathConfig} ')
            sys.exit(20)

    if not os.path.isfile(appConfigFile):
        try:
            # copy configuration file from GitHub clone
            shutil.copy(appConfigFile, pathConfig)
        except:
            colorFail('[e]', f'Unable to copy configuration file from the GitHub clone: {appConfigFile}')
            sys.exit(21)
    
    config = configparser.ConfigParser()
    config.read(appConfigFile)

    # Get Settings.
    # Future: Replace variables throughout with the direct appConfig reference
    studentName = config['Settings']['your_name']
    studentEmail = config['Settings']['email']
    
    # If blank settings exist prompt for value
    # Prompt to reuse or enter new psuedonym
    if '' == studentName:
        colorNotice(f'What is your name?')
        studentName = (str(input('>  ')))
        config['Settings']['your_name'] = studentName
    if '' == studentEmail:
        colorNotice(f'What is your email?   Enter to skip.\nThis is used to create a directory for your personal TTP collection.')
        studentEmail = (str(input('>  ')))
        config['Settings']['email'] = studentEmail

    if not os.path.exists(config['Paths']['pathwork']):
        os.umask(0o007)
        colorDebug(f"pathConfig does not exist: {config['Paths']['pathwork']}")
        try:
            os.mkdir(config['Paths']['pathwork'], 0o770)
        except:
            colorFail('[e]', f"Unable to create directory for AutoRpt settings: {config['Paths']['pathwork']} ")
            sys.exit(22)

    # Write new config.toml
    saveConfig(config)
    msg = "# WARNING !\n"
    msg += "Use private Git project repositories.\n\n"
    msg += "Do not use public Git projects.\n"
    msg += "Public projects may violate terms of service, non-disclosure agreements, or leak proprietary information.\n\n"
    msg += "\n"
    readme = f"{config['Paths']['pathwork']}/README.md"
    if not os.path.isfile(readme):
        with open(readme, 'w', encoding='utf-8') as f:
            f.write(msg)
            f.close()
        
    # If team notes directory does not exist, create it.
    # This is for your Team TTP collection or company specific documentation.
    ttp_notes_dir = f"{config['Paths']['pathwork']}/All-TTPs"
    if not os.path.exists(ttp_notes_dir):
        # If umask is not set, incorrect permissions will be assigned on mkdir
        os.umask(0o007)
        try:
            os.mkdir(ttp_notes_dir, 0o770)
        except:
            colorFail('[e]', f'Unable to create directory: {ttp_notes_dir} ')
            sys.exit(23)

    return config

def loadSessionConfig(appSessionFile):
    """Read session engagement file contents"""
    if os.path.isfile(appSessionFile):
        config = configparser.ConfigParser()
        config.read(appSessionFile)
        return config
    else:
        colorNotice('The session file does not exist. It will be created on first use of startup.')

def saveConfig(appConfig):
    """Store to disk modified application configuration values"""
    with open(appConfigFile, 'w') as configFile:
        appConfig.write(configFile)

def saveEnagements():
    """Store to disk session engagement values"""
    with open(sessionFile, 'w') as configFile:
        session.write(configFile)

def debugConfig():
    """Depricated. Verify reading of the configuration files were successful"""
    colorVerification('autorpt_runfrom', autorpt_runfrom)
    colorVerification('main appConfig', appConfig.sections())
    for key in appConfig['Paths']:
        colorVerification('App working path for engagenments', appConfig['Paths'][key])
    colorVerification('main session', session.sections())
    colorVerification('Active Engagement', session['Current']['Active'])
    colorVerification('Active Engagement Details for name', activeSessionDetails)
    for key in session['Engagements']:
        colorVerification(f'Engagement: {key}', session['Engagements'][key].split(','))
    print('\n')
    colorVerification('supported filetypes', supported_filetypes)
    colorVerification('no templates', extentionsWithoutTemplate)
    colorVerification('target file', targetsFile)
    colorVerification('port XLSX', portsSpreadsheet)
    colorVerification('vuln CSV', vulnsCsv)
    colorVerification('sitrep Log', sitrepLog)

def whathaveidone():
    """Summary analysis of session engagements."""
    # Super secret functionality.  jk.
    df = pd.DataFrame({})
    status = []
    types = []
    platforms = []

    for key in session.sections():
        if key not in ['DEFAULT', 'Current']:
            # to dataframe for analysis
            status.append(session[key]['status'])
            types.append(session[key]['type'])
            platforms.append(session[key]['platform'])
    
    new_row = {'STATUS': status,
            'TYPE': types, 
            'PLATFORM': platforms}
    df = pd.DataFrame(new_row)
    colorHeader("Activity Summary")
    pivot = df.pivot_table(index=['TYPE', 'STATUS'], values=['PLATFORM'], aggfunc='count').rename(columns={'PLATFORM': 'COUNT'})
    colorNotice(pivot)

    colorNotice(f'\n{term.bold}Total number of enagements: {df.shape[0]}{term.normal}\n') # row count
    
    colorSubHeading("Count of engagements by Status")
    colorNotice(df.STATUS.value_counts().to_string(index=True))
    
    colorSubHeading("\nCount of engagements by Type")
    colorNotice(df.TYPE.value_counts().to_string(index=True))
    
    colorSubHeading("\nCount of engagements by Platform")
    colorNotice(df.PLATFORM.value_counts().to_string(index=True))

    colorSubHeading("\nDetails")
    for key in session.sections():
        if key not in ['DEFAULT', 'Current']:
            # Either keep as is, a simple print, or add to dataframe and sort by status.
            # Currently sorted by age, oldest to most recent.
            print(f"{session[key]['status']:15}\t{key}")


# DisplayablePath from: 
# https://stackoverflow.com/questions/9727673/list-directory-tree-structure-in-python
class DisplayablePath(object):
    """Display a directory tree."""
    display_filename_prefix_middle = '├──'
    display_filename_prefix_last = '└──'
    display_parent_prefix_middle = '    '
    display_parent_prefix_last = '│   '

    def __init__(self, path, parent_path, is_last):
        self.path = Path(str(path))
        self.parent = parent_path
        self.is_last = is_last
        if self.parent:
            self.depth = self.parent.depth + 1
        else:
            self.depth = 0

    @property
    def displayname(self):
        if self.path.is_dir():
            return self.path.name + '/'
        return self.path.name

    @classmethod
    def make_tree(cls, root, parent=None, is_last=False, criteria=None):
        root = Path(str(root))
        criteria = criteria or cls._default_criteria

        displayable_root = cls(root, parent, is_last)
        yield displayable_root

        children = sorted(list(path
                               for path in root.iterdir()
                               if criteria(path)),
                          key=lambda s: str(s).lower())
        count = 1
        for path in children:
            is_last = count == len(children)
            if path.is_dir():
                yield from cls.make_tree(path,
                                         parent=displayable_root,
                                         is_last=is_last,
                                         criteria=criteria)
            else:
                yield cls(path, displayable_root, is_last)
            count += 1

    @classmethod
    def _default_criteria(cls, path):
        return True

    @property
    def displayname(self):
        if self.path.is_dir():
            return self.path.name + '/'
        return self.path.name

    def displayable(self):
        if self.parent is None:
            return self.displayname

        _filename_prefix = (self.display_filename_prefix_last
                            if self.is_last
                            else self.display_filename_prefix_middle)

        parts = ['{!s} {!s}'.format(_filename_prefix,
                                    self.displayname)]

        parent = self.parent
        while parent and parent.parent is not None:
            parts.append(self.display_parent_prefix_middle
                         if parent.is_last
                         else self.display_parent_prefix_last)
            parent = parent.parent

        return ''.join(reversed(parts))

if __name__ == "__main__":
    # Define a terminal for color sugar
    term = blessings.Terminal(kind='xterm-256color')
    # Display pretty ASCII art
    banner()
    # Get the script home starting directory (eg. /opt/AutoRpt)
    autorpt_runfrom = os.path.dirname(os.path.realpath(__file__))
    # Directory for additional, supporting content.
    # Currently only the Mitre ATT&CK Framwork.
    pathIncludes: autorpt_runfrom + '/includes'
    # Path to store configuration settings and sessions
    pathConfig = os.path.expanduser("~/.config/AutoRpt")
    # Configuration settings
    appConfigFile = pathConfig + '/config.toml'
    # Load configuration settings
    appConfig = loadAppConfig(pathConfig, appConfigFile)
    # Engagement sessions
    sessionFile = pathConfig + '/' + appConfig['Files']['sessionFile']
    session = loadSessionConfig(sessionFile)
    # Details for the active engagement or blank if none
    try:
        activeSessionDetails = session['Engagements'][session['Current']['active']].split(',')
    except:
        activeSessionDetails = ''
    # Should be supportedFiletypes
    supported_filetypes = appConfig['Settings']['output_formats']
    # Exclude filetypes that break report creation with pandoc
    # File with list of target IP addresses
    targetsFile = appConfig['Files']['targetFile']
    # Spreadsheet of all ports per IP address in targets file
    portsSpreadsheet = appConfig['Files']['portFile']
    # Validated list of vulnerabilities
    vulnsCsv =  appConfig['Files']['vulnFile']
    # Situation report
    sitrepLog =  appConfig['Files']['sitrepFile']
    
    # Take action based on parameters
    if len(sys.argv) <= 1:
        mainMenu()
    else:
        params(sys.argv[1:])