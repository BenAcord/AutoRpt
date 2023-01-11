#!/usr/bin/python3
"""
autorpt.py
Enforce consistent, dependable workflow for engagement note-taking and report writing.
"""

import os
import sys
from glob import glob
import csv
import re
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from cvss import CVSS3
import autorpt as core
import autorpt_cfg as cfg
import autorpt_pretty as out

def score_cvss3_basic():
    """ Rating calculation for the CVSS3 basic metric group. """
    basic_metric_value = ""

    out.color_subheading('Exploit')
    basic_metric_value += get_cvss_metric_value(
        {
            'Network': 'AV:N',
            'Adjacent Network': 'AV:A',
            'Local': 'AV:L',
            'Physical': 'AV:P'
        },
        "Attack Vector"
    )
    basic_metric_value += '/'

    basic_metric_value += get_cvss_metric_value(
        {
            'Low': 'AC:L',
            'High': 'AC:H'
        },
        "Attack Complexity"
    )
    basic_metric_value += '/'

    basic_metric_value += get_cvss_metric_value(
        {
            'None': 'PR:N',
            'Low': 'PR:L',
            'High': 'PR:H'
        },
        "Privileges Required"
    )
    basic_metric_value += '/'

    basic_metric_value += get_cvss_metric_value(
        {
            'None': 'UI:N',
            'Required': 'UI:R'
        },
        "User Interaction"
    )
    basic_metric_value += '/'

    basic_metric_value += get_cvss_metric_value(
        {
            'Unchanged': 'S:U',
            'Changed': 'S:C'
        },
        "Scope"
    )
    basic_metric_value += '/'

    out.color_subheading('Impact')
    basic_metric_value += get_cvss_metric_value(
        {
            'None': 'C:N',
            'Low': 'C:L',
            'High': 'C:H'
        },
        "Confidentiality Impact"
    )
    basic_metric_value += '/'

    basic_metric_value += get_cvss_metric_value(
        {
            'None': 'I:N',
            'Low': 'I:L',
            'High': 'I:H'
        },
        "Integrity Impact"
    )
    basic_metric_value += '/'

    basic_metric_value += get_cvss_metric_value(
        {
            'None': 'A:N',
            'Low': 'A:L',
            'High': 'A:H'
        },
        "Availability Impact"
    )
    return basic_metric_value.upper()

def score_cvss3_temporal():
    """ Rating calculation for the CVSS3 temporal metric group. """
    temporal_metric_value = "/"

    temporal_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'E:X',
            'Unproven that exploit exists': 'E:U',
            'Proof of concept code': 'E:P',
            'Functional exploit exists': 'E:F',
            'High': 'E:H'
        },
        "Exploit Code Maturity"
    )
    temporal_metric_value += '/'

    temporal_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'RL:X',
            'Official fix': 'RL:O',
            'Temporary fix': 'RL:T',
            'Workaround': 'RL:W',
            'Unavailable': 'RL:U'
        },
        "Remediation Level"
    )
    temporal_metric_value += '/'

    temporal_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'RC:X',
            'Unknown': 'RC:U',
            'Reasonable': 'RC:R',
            'Confirmed': 'RC:C'
        },
        "Report Confidence"
    )
    return temporal_metric_value.upper()

def score_cvss3_environmental():
    """ Rating calculation for the CVSS3 environmental metric group. """
    envionmental_metric_value = "/"

    out.color_subheading('Exploit')
    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'MAV:X',
            'Network': 'MAV:N',
            'Adjacent Network': 'MAV:A',
            'Local': 'MAV:L',
            'Physical': 'MAV:P'
        },
        "Environmental Attack Vector"
    )
    envionmental_metric_value += '/'

    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'MAC:X',
            'Low': 'MAC:L',
            'High': 'MAC:H'
        },
        "Environmental Attack Complexity"
    )
    envionmental_metric_value += '/'

    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'MPR:X',
            'None': 'MPR:N',
            'Low': 'MPR:L',
            'High': 'MPR:H'
        },
        "Environmental Privileges Required"
    )
    envionmental_metric_value += '/'

    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'MUI:X',
            'None': 'MUI:N',
            'Required': 'MUI:R'
        },
        "Environmental User Interaction"
    )
    envionmental_metric_value += '/'

    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'MS:X',
            'Unchanged': 'MS:U',
            'Changed': 'MS:C'
        },
        "Environmental Scope"
    )
    envionmental_metric_value += '/'

    out.color_subheading('Impact')
    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'MC:X',
            'None': 'MC:N',
            'Low': 'MC:L',
            'High': 'MC:H'
        },
        "Environmental Confidentiality Impact"
    )
    envionmental_metric_value += '/'

    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'MI:X',
            'None': 'MI:N',
            'Low': 'MI:L',
            'High': 'MI:H'
        },
        "Environmental Integrity Impact"
    )
    envionmental_metric_value += '/'

    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'MA:X',
            'None': 'MA:N',
            'Low': 'MA:L',
            'High': 'MA:H'
        },
        "Environmental Availability Impact"
    )
    envionmental_metric_value += '/'

    out.color_subheading('Impact Subscore Modifiers')
    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'CR:X',
            'Low': 'CR:L',
            'Medium': 'CR:M',
            'High': 'CR:H'
        },
        "Environmental Confidentiality Requirement"
    )
    envionmental_metric_value += '/'

    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'IR:X',
            'Low': 'IR:L',
            'Medium': 'IR:M',
            'High': 'IR:H'
        },
        "Environmental Integrity Requirement"
    )
    envionmental_metric_value += '/'

    envionmental_metric_value += get_cvss_metric_value(
        {
            'Not Defined': 'AR:X',
            'Low': 'AR:L',
            'Medium': 'AR:M',
            'High': 'AR:H'
        },
        "Environmental Availability Requirement"
    )
    return envionmental_metric_value.upper()

def get_cvss3_score():
    """Menu prompting for the vulnerability CVSS scoring"""
    out.color_header("CVSS 3 Scoring")
    cvss_vector = ""
    print("Do you know the Overall CVSS v3 Score? [Y|N]")
    known_score_response = str(input(">  ")).upper()
    if "Y" == known_score_response:
        print("What is the score?")
        cvss_score = float(input(">  "))
        if cvss_score <= 3.9:
            cvss_severity = 'Low'
        elif 4.0 <= cvss_score <= 6.9:
            cvss_severity = 'Medium'
        elif 7.0 <= cvss_score <= 8.9:
            cvss_severity = 'High'
        elif 9.0 <= cvss_score <= 10.0:
            cvss_severity = 'Critical'
        else:
            out.color_notice('The score must be between 0.1 and 10.0.')
            get_cvss3_score()
        this_msg = (
            'If known, paste the CVSS Vector string here'
            'or hit Enter to skip (eg. AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:N'
        )
        print(this_msg)
        cvss_vector = str(input('>  ')).upper()
        if cvss_vector in ('NONE', ''):
            # Was: 'NONE' == cvss_vector or '' == cvss_vector:
            cvss_vector = ''
        else:
            cvss_vector = "CVSS:3.0/" + cvss_vector
        return_string = [cvss_severity, str(cvss_score), cvss_vector]
    else:
        # At a bare minimum the base score is needed.
        # For testing - AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:N'
        out.color_subheading('Base Score Metrics')
        cvss_vector = 'CVSS:3.0/'
        cvss_vector += score_cvss3_basic()

        print(
            "Do you want to include CVSS v3 Temporal metrics in the score "
            "(3 additional questions)? [Y|N]"
        )
        print("If not, the score can be calculated from what has already been entered.")
        user_input = str(input(">  ")).upper()
        if "Y" == user_input:
            out.color_subheading('Temporal Score Metrics')
            cvss_vector += score_cvss3_temporal()


        print(
            "Do you want to include CVSS v3 Environmental metrics in the score "
            "(11 additional questions)? [Y|N]"
        )
        print("If not, the score can be calculated from what has already been entered.")
        user_input = str(input(">  ")).upper()
        if "Y" == user_input:
            out.color_subheading('Environmental Score Metrics')
            cvss_vector += score_cvss3_environmental()

        #out.color_debug(f"check the vector [{cvss_vector}]")
        cvss_values = CVSS3(cvss_vector)
        #out.color_debug(f"CVSS3 Values: [{cvss_values}]")
        return_string = [
            str(cvss_values.severities()[2]),
            str(cvss_values.scores()[2]),
            cvss_vector
        ]
    return return_string

def get_cvss_metric_value(cvss_dictionary, metric_name):
    """Helper to display boilerplate prompts for values"""
    print("What is the " + metric_name + "?")
    for (i, opt) in enumerate(list(cvss_dictionary)):
        print("\t" + str(i) + ") " + opt)
    try:
        user_input = cvss_dictionary.get(list(cvss_dictionary)[int(input(" >  "))])
    except IndexError:
        out.color_fail('CVSS Value', 'That selection is not available in this listing.')
        sys.exit(34)
    return user_input

def get_mitre_attack():
    """Menu prompting to select MITRE ATT&CK tactic and technique"""
    tactic = ''
    technique = ''
    out.color_header("MITRE ATT&CK")
    path_includes = cfg.autorpt_runfrom + '/includes'
    csv_file_list = glob(path_includes + '/autorpt-*-attack.csv')

    i = 0
    out.color_notice('Which MITRE ATT&CK Framework applies?\nPress 99 for main menu.')
    for file in csv_file_list:
        out.color_menu_item(f'{i}. {file[30:-11]}')
        i = i + 1
    picker = int(input('>  '))
    if 99 == picker:
        core.main_menu()
    elif picker > (len(csv_file_list)):
        out.color_notice('Selection out of range')
        core.main_menu()
    else:
        file = csv_file_list[picker]

    this_dataframe = pd.read_csv(file, index_col=False, engine="python")
    

    # Get the tactic
    i = 0
    picker = 0
    tactics = this_dataframe.TACTIC.unique()
    out.color_notice('What is the Tactic?\nOr 99 to return to the ATT&CK menu.')
    for tactic in tactics:
        out.color_menu_item(f'{i}. {tactic}')
        i = i + 1
    picker = int(input('>  '))
    if 99 == picker:
        get_mitre_attack()
    elif picker > len(tactics):
        out.color_notice('Selection out of range.')
        core.main_menu()
    else:
        tactic = tactics[picker]

    # Get the technique
    i = 0
    picker = 0
    techniques = this_dataframe.query(f'TACTIC == "{tactic}"')[['TECHNIQUE']]
    out.color_notice('Pick a Technique?')
    for _, row in techniques.iterrows():
        # index variable replaced with _ as it is unused.
        out.color_menu_item(f"{str(i)}.  {str(row.TECHNIQUE)}")
        i = i + 1
    picker = int(input('>  '))
    technique = techniques.iloc[picker, 0]

    return [tactic, technique]

def get_nmap_file(target, ports_file):
    """A listing of known good nmap output files.
    In order: AutoRecon, nmapAutomator, and Reconnoitre."""

    nmap_file = ""
    nmap_file_list = ["_full_tcp_nmap.txt",
                 "_quick_tcp_nmap.txt",
                 f"Full_{target}.nmap",
                 f"{target}.quick.nmap",
                 f"{target}.nmap"]

    for name in nmap_file_list:
        for root, _, files in os.walk(cfg.get_active_path()):
            nmap_file = os.path.join(root, name)
            if name in files:
                get_nmap_file_contents(nmap_file, target, ports_file)
    if nmap_file == '':
        out.color_notice(f'\t\tExiting.  No nmap files found: [{nmap_file}]')
        sys.exit(35)

def get_nmap_file_contents(nmap_file, target, ports_file):
    """ Parse the contents of a discovered nmap file. Store results in XLSX. """

    this_dataframe = pd.DataFrame({})
    all_ports = pd.DataFrame({})
    ip_address = ''
    with open(nmap_file, 'r', encoding='utf-8', newline='') as nmap_file_reader:
        for line in nmap_file_reader.readlines():
            if re.match(r"^Nmap scan report for ", line):
                ip_address = line.strip().replace('Nmap scan report for ', '')
            elif re.match(r"^\d+.*$", line):
                # 0:port, 1:state, 2:service, 3:reason(skip), 4:version(glob)
                fields = line.strip().split()

                if fields[2] == 'unrecognized':
                    continue

                new_row = {
                    'IPADDRESS': ip_address,
                    'PORT': fields[0],
                    'STATE': fields[1],
                    'SERVICE': fields[2],
                    'VERSION': re.sub(r'\(.*\)', '', ' '.join(fields[4:]))
                }
                this_dataframe = this_dataframe.append(new_row, ignore_index = True)
                all_ports = all_ports.append(new_row, ignore_index = True)
        nmap_file_reader.close()
    # Provide CLI information.
    out.color_verify(target, f'Port Count: {str(len(this_dataframe.index))}')
    out.color_list(this_dataframe.to_markdown())
    # Create worksheet per target.
    if os.path.exists(ports_file):
        book = openpyxl.load_workbook(ports_file)
    else:
        book = openpyxl.Workbook()
    if 'Sheet' in book.sheetnames:
        del book['Sheet']
    try:
        sheet = book.create_sheet(target, 0)
        for df_row in dataframe_to_rows(this_dataframe, index=True, header=True):
            if str(df_row) != "[None]":
                sheet.append(df_row)
        book.save(ports_file)
    except (PermissionError, IOError):
        out.color_fail(
            'Ports Spreadsheet',
            'Cannot create or append worksheet to ports file.'
        )

def ports():
    """ Display the ports and replace the ports spreadsheet. """

    if os.path.isfile(f"{cfg.get_active_path()}/{cfg.targets_file}"):
        ports_file = f"{cfg.get_active_path()}/report/{cfg.ports_spreadsheet}"
        if os.path.isfile(ports_file):
            out.color_notice(f'Removing existing ports file: {ports_file}')
            os.remove(ports_file)
        # Look for nmap output files associated with each target IP address
        with open(
            f"{cfg.get_active_path()}/{cfg.targets_file}",
            'r',
            encoding='utf-8',
            newline=''
        ) as target_file_reader:
            targets = target_file_reader.readlines()
            for target in targets:
                target = target.strip()
                get_nmap_file(target, ports_file)
        # Update the engagement status
        active = cfg.session['Current']['active']
        cfg.session[active]['status'] = 'In-process'
        cfg.save_enagements()
